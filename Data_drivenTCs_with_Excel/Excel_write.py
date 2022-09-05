import openpyxl

path="D:\Empty1.xlsx"

workbook= openpyxl.load_workbook(path)
sheet= workbook.active

lister= ["A","B","C","D","E"]
for r in range(1,6):
    for c in range(1,6):
        if(r==c):
            sheet.cell(row=r,column=c).value= lister[r-1]
        else:
            sheet.cell(row=r,column=c).value = "-"

workbook.save(path)
