import openpyxl

path= "D:\Sample_list1.xlsx"

excel_file = openpyxl.load_workbook(path)
sheet= excel_file.active

rows= sheet.max_row
columns= sheet.max_column
print("ROWS: {}   COLUMNS:  {}".format(rows,columns))

for r in range(1,rows+1):
    for c in range(1, columns+1):
        print((sheet.cell(r,c).value),end="    ")
    print()

