import openpyxl

def getRowCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(file, sheetName,rowNo, columnNo):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return(sheet.cell(rowNo, columnNo).value)

def writeData(file, sheetName,rowNo, columnNo, data):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    sheet.cell(rowNo, columnNo).value = data
    workbook.save(file)
